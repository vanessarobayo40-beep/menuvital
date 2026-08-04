# -*- coding: utf-8 -*-
"""
MenúVital — Construye las recetas nuevas a partir de Recetas2026/ (material
fuente de Vanessa, ignorado por git) y las escribe en tools/recipes_build.json
para que tools/validate_recipes.py y tools/emit_php.py las procesen.

Uso:
    python build_recipes.py

Requiere que Recetas2026/ exista junto a este repo (no se versiona — ver
.gitignore). Si no está presente, el script no puede reconstruirse; por eso
este pipeline queda documentado y versionado aquí.
"""
import json
import os
import re
import shutil
import sys
import unicodedata

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from ingredient_parse import parse_ingredient_line, is_junk_line
from nutrition_calc import sum_macros
from nutrition_flags import is_vegetarian, is_gluten_free, is_economical
from normalize import normalize_ingredient
from unit_weights import WHOLE_UNIT_GRAMS
from nutrition_flags import ANIMAL_PROTEIN_KEYS

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, 'Recetas2026')
RECIPES_PHP = os.path.join(ROOT, 'database', 'recipes_data.php')
IMG_DST = os.path.join(ROOT, 'assets', 'img', 'recetas')
OUT_JSON = os.path.join(os.path.dirname(__file__), 'recipes_build.json')

MEAL_MAP = {'desayuno': 'desayuno', 'almuerzo': 'almuerzo', 'cena': 'cena', 'snack': 'snack'}

# Basura sistemática de la extracción de PDF (ver informe de exploración):
# ligaduras rotas ("Dizcultad"=Dificultad, "muxins"=muffins) y colas de página.
_DAMAGE_RE = re.compile(
    r'Dizcultad|muxins|zrme|Verizca|Page \d+|Recetario Pro|&#\d+;|AirFryer: Cocina')
_BAD_NAME_START = re.compile(r'^[^A-Za-zÁÉÍÓÚÑáéíóúñ0-9]')
# Señal de que dos columnas del PDF quedaron entrelazadas palabra por palabra:
# un verbo de instrucción en mayúscula que NO está al inicio de una oración
# ("...con papel Corta el tofu...") — en texto sano esos verbos solo abren
# una frase nueva después de punto, nunca aparecen sueltos a mitad de frase.
_INTERLEAVED_RE = re.compile(
    r'[a-záéíóúñ,]\s+(Corta|Agrega|Mezcla|Sofríe|Cocina|Retira|Sirve|Coloca|'
    r'Precalienta|Incorpora|Añade|Deja|Vierte|Espolvorea|Pela|Licúa|Hornea|'
    r'Sazona|Calienta|Rocía|Forma|Extiende|Revuelve|Tapa|Escurre)\b')
_TAIL_JUNK = re.compile(
    r'\s+(INGREDIENTES|VALOR N.*|Page \d+.*)$', re.IGNORECASE)


def norm_name(s: str) -> str:
    n = unicodedata.normalize('NFKD', str(s).lower())
    n = n.encode('ascii', 'ignore').decode()
    return re.sub(r'[^a-z0-9 ]', '', n).strip()


def load_existing_names():
    php = open(RECIPES_PHP, encoding='utf-8').read()
    return {norm_name(m.group(1)) for m in re.finditer(r"^\['([^']*)'", php, re.M)}


def clean_name(raw: str) -> str:
    name = str(raw).strip()
    name = _BAD_NAME_START.sub('', name).strip()
    name = _TAIL_JUNK.sub('', name).strip()
    name = re.sub(r'\s+', ' ', name)
    if len(name) > 150:
        name = name[:150].rsplit(' ', 1)[0]
    # Capitaliza solo la primera letra si vino toda en mayúsculas/minúsculas.
    if name and (name.isupper() or name.islower()):
        name = name[0].upper() + name[1:]
    return name


def parse_time_min(text, prep_text, meal_type):
    for src in (text, prep_text):
        if src:
            m = re.search(r'(\d+)\s*-?\s*(\d+)?\s*min', str(src))
            if m:
                a = int(m.group(1))
                b = int(m.group(2)) if m.group(2) else a
                return round((a + b) / 2)
    return {'desayuno': 10, 'snack': 10, 'almuerzo': 25, 'cena': 20}.get(meal_type, 20)


def split_steps(prep_text: str):
    if not prep_text:
        return []
    raw_lines = [l.strip() for l in str(prep_text).split('\n') if l.strip()]
    if len(raw_lines) < 2:
        # Todo en una sola línea: separar por "N. " (numeración).
        raw_lines = re.split(r'(?<=[.\)])\s*(?=\d+\.\s)', str(prep_text))
    steps = []
    for l in raw_lines:
        l = re.sub(r'^\d+[\.\)]\s*', '', l.strip())
        l = _DAMAGE_RE.sub('', l).strip()
        if len(l) < 8:
            continue
        if len(l) > 300:
            l = l[:297].rsplit(' ', 1)[0] + '...'
        steps.append(l)
    return steps


def infer_servings(parsed_ings):
    """
    Cuando la fuente no dice cuántas porciones rinde la receta, se infiere a
    partir de la masa de proteína principal: una porción de comida para un
    adulto trae entre 120 y 200 g de carne/pescado cocido (GABA/ICBF). Si la
    receta trae, p. ej., 700 g de alitas o "2 pechugas" enteras, es una
    receta para 2-4 personas, no para 1 — evita kcal absurdas por porción.
    """
    protein_grams = sum(
        p['grams'] for p in parsed_ings
        if p.get('match_name') in ANIMAL_PROTEIN_KEYS and p.get('match_name') not in
        {'huevo', 'clara de huevo', 'yema de huevo'}
    )
    if protein_grams <= 0:
        return 1
    servings = round(protein_grams / 180)
    return max(1, min(4, servings))


def parse_ingredients_cell(cell_text):
    """Devuelve (display_list, parsed_list) — display_list es ['nombre|qty', ...]."""
    if not cell_text:
        return [], []
    lines = [l.strip(' -\t*') for l in str(cell_text).split('\n') if l.strip(' -\t*')]
    display, parsed = [], []
    for raw in lines:
        if is_junk_line(raw):
            continue
        p = parse_ingredient_line(raw)
        if not p or not p['display_name']:
            continue
        name = p['display_name'][:60]
        qty = p['display_qty']
        display.append(f"{name}|{qty}" if qty and qty != 'al gusto' else f"{name}|al gusto")
        parsed.append(p)
    return display, parsed


def derive_tags(kcal, protein, time_min, match_names, is_colombian):
    tags = []
    if is_colombian:
        tags.append('tradicional')
    if protein >= 25 or (kcal > 0 and (protein * 4 / kcal) >= 0.30):
        tags.append('alto en proteína')
    if kcal <= 350 and 'ligero' not in tags:
        tags.append('ligero')
    if time_min <= 20:
        tags.append('rápido')
    if is_vegetarian(match_names):
        tags.append('vegetariano')
    if is_gluten_free(match_names):
        tags.append('sin gluten')
    if is_economical(match_names):
        tags.append('económico')
    # Únicos, preservando orden, tope de 4.
    seen, out = set(), []
    for t in tags:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out[:4]


FOLDER_CONFIG = {
    # carpeta: servings_forzado (None = inferir por masa de proteína, ver infer_servings)
    'metodo_airfryer': None,
    'bowls_express': None,
    'snacks_airfryer': None,
    'barras_airfryer': 8,   # lote de barras horneadas -> se reparte en 8 unidades
    'batidos_bono': None,
    'tortas': 8,            # torta/pan entero -> se reparte en 8 porciones (rebanadas)
    'air_fryer': None,
    'desayunos_mundo': None,
}

LEFTOVER_FOLDERS = ['250recetas', '120recetas', 'antiinflamatorias', 'diabeticos', 'higadograso', 'desayunos']


def find_col(hdr, prefix):
    for i, h in enumerate(hdr):
        if str(h).strip().lower().startswith(prefix):
            return i
    return None


def build_from_row(name_raw, tipo_raw, img_raw, ing_raw, prep_raw, time_raw, servings_hint,
                    source_tag, images_dir, existing_names, seen_names, warnings):
    name = clean_name(name_raw)
    if not name or norm_name(name) in existing_names or norm_name(name) in seen_names:
        return None
    meal_type = MEAL_MAP.get(str(tipo_raw or '').strip().lower())
    if not meal_type:
        return None

    display_ings, parsed_ings = parse_ingredients_cell(ing_raw)
    if not display_ings:
        return None
    steps = split_steps(prep_raw)
    if not steps:
        return None

    time_min = parse_time_min(time_raw, prep_raw, meal_type)
    servings = servings_hint or infer_servings(parsed_ings)
    macros, stats = sum_macros(parsed_ings, servings)
    if macros['kcal'] <= 0:
        return None

    match_names = [p['match_name'] for p in parsed_ings]
    tags = derive_tags(macros['kcal'], macros['protein'], time_min, match_names,
                        is_colombian=False)

    image_url = None
    if img_raw and str(img_raw).strip().lower() != 'none':
        fname = str(img_raw).strip()
        src_path = os.path.join(images_dir, fname)
        if os.path.isfile(src_path):
            dst_path = os.path.join(IMG_DST, fname)
            if not os.path.isfile(dst_path):
                shutil.copy2(src_path, dst_path)
            image_url = f'/assets/img/recetas/{fname}'

    seen_names.add(norm_name(name))
    confidence = 'alta' if stats['match_ratio'] >= 0.7 and stats['unquantified_count'] <= len(parsed_ings) * 0.4 else 'revisar'

    return {
        'name': name, 'meal_type': meal_type, 'ingredients': display_ings, 'steps': steps,
        'tags': tags, 'kcal': macros['kcal'], 'protein': macros['protein'],
        'time_min': time_min, 'carbs': macros['carbs'], 'fat': macros['fat'],
        'sugar': macros['sugar'], 'fiber': macros['fiber'], 'image_url': image_url,
        '_source': source_tag, '_confidence': confidence, '_stats': stats,
    }


def process_clean_folder(folder, existing_names, seen_names, warnings):
    path = os.path.join(SRC, folder, 'Recetas.xlsx')
    images_dir = os.path.join(SRC, folder, 'images')
    ws = openpyxl.load_workbook(path, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) for h in rows[0]]
    ci_name, ci_img, ci_tipo = 1, 2, 3
    ci_time = find_col(hdr, 'tiempo')
    ci_serv = find_col(hdr, 'porcion')
    ci_ing = find_col(hdr, 'ingrediente')
    ci_prep = find_col(hdr, 'preparaci')

    out = []
    for r in rows[1:]:
        if r[ci_name] is None:
            continue
        serv_hint = None
        if ci_serv is not None and r[ci_serv]:
            m = re.search(r'\d+', str(r[ci_serv]))
            if m:
                serv_hint = int(m.group())
        rec = build_from_row(
            r[ci_name], r[ci_tipo], r[ci_img], r[ci_ing], r[ci_prep],
            r[ci_time] if ci_time is not None else None,
            serv_hint or FOLDER_CONFIG[folder],
            f'Recetas2026/{folder}', images_dir, existing_names, seen_names, warnings,
        )
        if rec:
            out.append(rec)
    return out


def process_recetario_pro(existing_names, seen_names, warnings):
    folder = 'recetario_pro_airfryer'
    path = os.path.join(SRC, folder, 'Recetas.xlsx')
    images_dir = os.path.join(SRC, folder, 'images')
    ws = openpyxl.load_workbook(path, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for r in rows[1:]:
        if r[1] is None:
            continue
        name_raw, img_raw, tipo_raw, time_raw = r[1], r[2], r[3], r[4]
        serv_raw, ing_raw, prep_raw = r[5], r[6], r[7]
        nm, prep = str(name_raw).strip(), str(prep_raw or '').strip()
        ing = str(ing_raw or '').strip()
        ok = (ing and ing != 'None' and not _DAMAGE_RE.search(prep) and not _DAMAGE_RE.search(nm)
              and not _DAMAGE_RE.search(ing) and not _BAD_NAME_START.search(nm)
              and len(prep) > 80 and not _INTERLEAVED_RE.search(prep))
        if not ok:
            warnings.append(f'[recetario_pro_airfryer] descartada (dañada por extracción PDF): {nm!r}')
            continue
        serv_hint = None
        if serv_raw:
            m = re.search(r'\d+', str(serv_raw))
            if m:
                serv_hint = int(m.group())
        rec = build_from_row(
            name_raw, tipo_raw, img_raw, ing_raw, prep_raw, time_raw, serv_hint,
            f'Recetas2026/{folder}', images_dir, existing_names, seen_names, warnings,
        )
        if rec:
            out.append(rec)
    return out


def process_leftovers(existing_names, seen_names, warnings):
    out = []
    for folder in LEFTOVER_FOLDERS:
        path = os.path.join(SRC, folder, 'Recetas.xlsx')
        if not os.path.isfile(path):
            continue
        images_dir = os.path.join(SRC, folder, 'images')
        ws = openpyxl.load_workbook(path, data_only=True).active
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(h) for h in rows[0]]
        ci_ing = find_col(hdr, 'ingrediente')
        ci_prep = find_col(hdr, 'preparaci')
        ci_time = find_col(hdr, 'tiempo')
        if ci_ing is None or ci_prep is None:
            continue
        for r in rows[1:]:
            if r[1] is None:
                continue
            if norm_name(r[1]) in existing_names:
                continue  # ya existe en el recetario actual, no es "sobra"
            ing_txt = str(r[ci_ing] or '')
            if len(ing_txt) <= 25 or 'ver preparaci' in ing_txt.lower():
                continue  # sin ingredientes reales, no se puede reconstruir
            img_raw = r[2] if len(r) > 2 else None
            tipo_raw = r[3] if len(r) > 3 else None
            rec = build_from_row(
                r[1], tipo_raw, img_raw, r[ci_ing], r[ci_prep],
                r[ci_time] if ci_time is not None else None, None,
                f'Recetas2026/{folder} (sobra)', images_dir, existing_names, seen_names, warnings,
            )
            if rec:
                out.append(rec)
    return out


def main():
    if not os.path.isdir(SRC):
        print(f'ERROR: no existe {SRC} — este build solo corre en el equipo con el material fuente.')
        sys.exit(1)

    existing_names = load_existing_names()
    seen_names = set()
    warnings = []

    all_recipes = []
    for folder in FOLDER_CONFIG:
        recs = process_clean_folder(folder, existing_names, seen_names, warnings)
        print(f'{folder}: {len(recs)} recetas')
        all_recipes += recs

    # NOTA: recetario_pro_airfryer se descarta por completo. El primer filtro
    # (heurística de nombre/kcal) dejaba pasar 94 filas; al revisar una
    # muestra a mano se vio que muchas seguían con instrucciones entrelazadas
    # del PDF a dos columnas ("en cubos pequeños." como paso suelto,
    # "Corta el pepino" colado como ingrediente). Un segundo filtro más
    # estricto (_INTERLEAVED_RE) bajó las sanas de 94 a 42, y una revisión
    # manual de esas 42 mostró que casi todas seguían teniendo el mismo
    # problema en menor grado. No es viable garantizar "cero riesgo de pasos
    # ilegibles" para esta carpeta con reglas automáticas — se necesitaría
    # volver a extraer del PDF original. Queda fuera de este build; ver
    # informe_validacion.md y el mensaje final para Vanessa.
    # recs = process_recetario_pro(existing_names, seen_names, warnings)
    # print(f'recetario_pro_airfryer (solo filas sanas): {len(recs)} recetas')
    # all_recipes += recs

    recs = process_leftovers(existing_names, seen_names, warnings)
    print(f'sobras de carpetas viejas: {len(recs)} recetas')
    all_recipes += recs

    by_type = {}
    for r in all_recipes:
        by_type[r['meal_type']] = by_type.get(r['meal_type'], 0) + 1
    print('\nTotal nuevas (sin colombianas):', len(all_recipes))
    print('Por tipo:', by_type)

    low_conf = [r for r in all_recipes if r['_confidence'] == 'revisar']
    print(f'Confianza baja (revisar manualmente): {len(low_conf)}')

    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(all_recipes, f, ensure_ascii=False, indent=1)
    print(f'\nEscrito: {OUT_JSON}')

    if warnings:
        warn_path = os.path.join(os.path.dirname(__file__), '_build_warnings.txt')
        with open(warn_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(warnings))
        print(f'Avisos ({len(warnings)}): {warn_path}')


if __name__ == '__main__':
    main()
